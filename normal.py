from os import sep
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from pandas.core.indexes.base import Index   

'''CON CSV donde sacare los datos par amandar el valor 1
#Lectura del archivo CSV
arch_csv = pd.read_csv('1.csv', sep =',')
#jsCSV = arch_csv[['Participants']]
datosCVS = arch_csv[['Participants']]
print(df.loc['Participants'])
#print(type(datosCVS))
for i in datosCVS:
    print(f"DATO: {i}")
    pass '''
'''Ahora con openpyxl hacemos esto'''    
arch_csv = pd.read_csv('1.csv', sep =',')
datosCVS = arch_csv[['Participants']]
datosCVS.to_excel('1.xlsx')
wb = load_workbook(filename = '1.xlsx',data_only=True)
wb2 = wb['Sheet1']
# data = wb2["B2"].value
# data2 = data.split(" ")

# range_cell = wb2['B2':'B47']

# cell_range = wb['B2':'B47']
# print(wb.sheetnames)
# print(range_cell)
                #values_only=True son para los valores que tiene dentro de la celda, dentro
for col in wb2.iter_cols(min_row=2, max_col=2, max_row=47, values_only=True):
    # print(col)
    col = list(col)
    # print(col)
    # for cell in col:
    #     print(cell)

for nombre in col:
    arr_nombre = nombre.split(" ")
    print(arr_nombre)

#datosCVS.to_excel('1.xlsx')
'''esta lines convierte la consulta del CVS a json'''
#jsCSV.to_json('jsCSV.json')
#listaCSV = []
#listaCSV.append(arch_csv[['Participants']])
#print(listaCSV)
#print(arch_csv[['Participants']])
#print(arch_csv[['Participants']]

'''CON XLSX es a donde voy a mandar el 1 cuando Participants sea igual entre B19;B66
arch_excel = pd.read_excel('2.xlsx')

en esta linea se solicitam los datos que se requieren para validar los datos del CVS con el XLSX
jsXLSX = arch_excel[['CONTROL DE ASISTENCIA']].iloc[17:66]
#print(jsXLSX) '''